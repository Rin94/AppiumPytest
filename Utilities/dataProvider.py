import openpyxl
import os
def get_data(sheetName):
    workbook = openpyxl.load_workbook(os.getcwd()+"/../excel/testData.xlsx")
    sheet = workbook[sheetName]
    mainList = []
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    for i in range(2, totalrows+1):
        dataList = []
        for j in range(1,totalcols+1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList